#!/usr/bin/env python
# coding: utf-8

# In[1]:




from pathlib import Path
import numpy as np
import pandas as pd
from fpdf import FPDF
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.ticker import ScalarFormatter
import openpyxl

xlsx_file = Path( 'Dummy Data.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file) 

# Read the active sheet:
sheet = wb_obj.active


# In[4]:


pdf = FPDF(orientation = 'P', unit = 'mm', format = 'A4')
k=3
reg = (int(sheet.cell(row=3, column=6).value))
for j in range (1,6):
    pdf.add_page()
    pdf.set_font('Arial', 'B', 11)
    pdf.set_text_color(0,0,0)
    pdf.image('Question.png', x = 0, y = 0, w = 210, h = 297)
    for_total = 0
    total = 0
    reg+=((28)*(j-1))
    pdf.text(75,17 ,(str(reg)))
    pdf.text(75,23 , (str(sheet.cell(row=k, column=3).value)))
    pdf.text(75,28 , (str(sheet.cell(row=k, column=4).value)))
    pdf.text(75,33 , ((str(sheet.cell(row=k, column=3).value))+" "+(str(sheet.cell(row=k, column=4).value))))
    pdf.image(''+(str(sheet.cell(row=k, column=3).value))+' '+(str(sheet.cell(row=k, column=4).value))+'.jpg', x = 120, y = 17, w = 50, h = 50)
    pdf.image('logo.png', x = 10, y = 90, w = 50, h = 50)
    pdf.text(75,40 , (str(sheet.cell(row=k, column=9).value)))
    pdf.text(75,46 , (str(sheet.cell(row=k, column=10).value)))
    pdf.text(75,51 , (str(sheet.cell(row=k, column=8).value)))
    pdf.text(75,57 , (str(sheet.cell(row=k, column=7).value)))
    pdf.text(75,62 , (str(sheet.cell(row=k, column=11).value)))
    pdf.text(75,69 , (str(sheet.cell(row=k, column=13).value)))
    pdf.text(100,78 , (str(sheet.cell(row=k, column=2).value)))
    pdf.text(115,75 , 'Date and Time of Exam')
    pdf.text(170,75 , (str(sheet.cell(row=k, column=12).value)))
    pdf.set_font('Arial', 'B', 11)
    pdf.text(80,280 , (str(sheet.cell(row=k, column=20).value)))
    for i in range(3,28):
        pdf.text(115,98+(5.6*(i-3)) , (str(sheet.cell(row=k, column=15).value)))
        pdf.text(132,98+(5.6*(i-3)) , (str(sheet.cell(row=k, column=16).value)))
        crtc = (str(sheet.cell(row=k, column=16).value))
        if (str(sheet.cell(row=k, column=15).value)) == (str(sheet.cell(row=k, column=16).value)):
            pdf.set_text_color(0,255,0)
            pdf.text(150,98+(5.6*(i-3)) , "C")
        elif (str(sheet.cell(row=k, column=15).value)) == 'None':
            pdf.set_text_color(0,0,255)
            pdf.text(150,98+(5.6*(i-3)) , "N")
        elif (str(sheet.cell(row=k, column=15).value)) != (str(sheet.cell(row=k, column=16).value)):
            pdf.set_text_color(255,0,0)
            pdf.text(150,98+(5.6*(i-3)) , "W")
        pdf.set_text_color(0,0,0)
        pdf.text(168,98+(5.6*(i-3)) , (str(sheet.cell(row=k, column=18).value)))
        for_total+= (int(sheet.cell(row=k, column=18).value))
        if (str(sheet.cell(row=k, column=15).value)) == (str(sheet.cell(row=k, column=16).value)):
            pdf.text(186,98+(5.6*(i-3)) , (str(sheet.cell(row=k, column=18).value)))
            total+=  (int(sheet.cell(row=k, column=18).value))
        elif (str(sheet.cell(row=k, column=15).value)) != (str(sheet.cell(row=k, column=16).value)):
            pdf.text(186,98+(5.6*(i-3)) , (str(0)))
        k+=1

    
    Scored = 'right', 'wrong'    
    marks = [total, for_total-total]    
    explode = (0.1, 0)  # it "explode" the 1st slice     
    
    fig1, ax1 = plt.subplots()    
    ax1.pie(marks, explode=explode, labels=Scored, autopct='%1.1f%%',    
        shadow=True, startangle=90)    
    ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.    
    

    plt.savefig('a.png')
    pdf.image('a.png', x = -20, y = 140, w = 100, h = 100)
        
pdf.output ('python.pdf')


# In[ ]:





# In[ ]:




